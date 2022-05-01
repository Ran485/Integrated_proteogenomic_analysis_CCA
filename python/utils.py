# here put the import lib
import pandas as pd
import numpy as np
import seaborn as sn
import os
import datetime
from sklearn.impute import KNNImputer
from sklearn.preprocessing import MinMaxScaler
import plotly.graph_objs as go
import plotly.express as px
import matplotlib.pyplot as plt
# python matplotlib PDF 不断字
import matplotlib as mpl
from matplotlib_venn import venn2, venn2_circles
from matplotlib_venn import venn3, venn3_circles
mpl.rcParams['pdf.fonttype'] = 42
mpl.rc('font',family='Arial',size=12)
from openpyxl import load_workbook


def paste0(a, b):
    # Concatenating multiple strings with a single string
    return a + b


def create_output_dir(filename = None, creat_time_subdir = False):
    """create the output directory

    Args:
    -------------
        filename ([str]): A given filename  

        creat_time_subdir (bool, optional): 
                        creat 2021-11-12 subdirectory,defaults to False.
    Returns:  
    -------------
        output_dir [str]: the output directory
    """
    root_dir = './'
    out_path = root_dir + filename
    if not os.path.isdir(out_path):  # in case root_dir doesn't exist
        os.makedirs(out_path)
        print("Successfully created output subdir: {}".format(out_path))
        if creat_time_subdir:
            date_string = datetime.now().strftime("%Y_%m_%d")
            out_path = os.path.join(out_path, date_string)
            if not os.path.isdir(out_path):
                os.mkdir(out_path)
                print("Successfully created output subdir: {}".format(out_path))
    else:
        print("The current path: {} already exist".format(out_path))
    return out_path + '/'


def read_specific_sheet(path, sheet_name):
    """read specific sheet from excel file
    """
    workbook = load_workbook(filename=path)
    names = workbook.sheetnames
    print(names)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"The title of the Worksheet is: {sheet.title}")
        print(f"Cells that contain data: {sheet.calculate_dimension()}")
        return sheet


def save_multisheet_excel(append_file=None,
                          input_file=None,
                          filetype=None,
                          sheet_name="table name to be saved"):
    """
    read table and write it to the multisheet excel file
    """

    # excel file use 'pd.read_excel', csv/txt file use 'pd.read_csv'
    if isinstance(input_file, str):
        if filetype == 'excel':
            df = pd.read_excel(input_file)
        elif filetype == 'csv':
            df = pd.read_csv(input_file)
        else:
            df = pd.read_csv(input_file, sep="\t")
    else:
        df = input_file
        # df.reset_index(inplace=True)

    # 对 df 的数据进行相应的处理, 一般操作的话将其注释掉
    # df.replace(np.nan, 0, inplace=True)

    # return df
    with pd.ExcelWriter(append_file, mode="a", engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)


# if __name__ == '__main__':
#        input_file = r"/Users/ranpeng/Desktop/CCA/CCA_mapping_data/phosphoprotein_abudance_FOT(2021-2-3).csv"
#        append_file = r"/Users/ranpeng/Desktop/CCA/Supplementary_Tables/Table1.xlsx"
#        save_multisheet_excel(append_file, input_file, filetype = 'csv', sheet_name="phosphoprotein_abudance")
